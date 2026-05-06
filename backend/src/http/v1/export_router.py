"""
ROUTER: export_router
PURPOSE: Export company and personnel data to CSV/Excel
ENCODING: UTF-8 WITHOUT BOM
"""

from fastapi import APIRouter, HTTPException, Depends, Header
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import List, Optional, Dict, Any
import csv
import io
from datetime import datetime
from sqlalchemy.orm import Session

from src.models.user import User
from src.db.base import get_db
from src.services.auth_service import get_current_user_dependency, get_current_user_impl

router = APIRouter(prefix="/export", tags=["export"])


***REMOVED*** Helper function to create dependency with proper header injection
def get_current_user_from_header(
    authorization: str = Header(None, alias="Authorization"),
    db: Session = Depends(get_db)
):
    """Helper to inject authorization header into get_current_user_dependency"""
    return get_current_user_impl(authorization=authorization, db=db)


class ExportRequest(BaseModel):
    data_type: str  ***REMOVED*** "companies", "personnel", "both"
    format: str = "csv"  ***REMOVED*** "csv" or "excel"
    filters: Optional[Dict[str, Any]] = None


@router.post("/companies")
async def export_companies(
    request: ExportRequest,
    user: User = Depends(get_current_user_from_header)
):
    """
    Export company data to CSV or Excel.
    Protected route - requires authentication.
    """
    ***REMOVED*** Mock company data - in production, get from database
    companies = [
        {
            "company_id": "comp-001",
            "name": "Grand Hotel Istanbul",
            "address": "Taksim, Istanbul",
            "website": "www.grandhotel.com",
            "phone": "+90 212 123 4567",
            "email": "info@grandhotel.com",
            "source": "Official Source",
            "status": "Verified",
            "created_at": datetime.utcnow().isoformat()
        },
        {
            "company_id": "comp-002",
            "name": "Bakery Co.",
            "address": "Kadıköy, Istanbul",
            "website": "www.bakeryco.com",
            "phone": "+90 216 234 5678",
            "email": "contact@bakeryco.com",
            "source": "Internet",
            "status": "Verified",
            "created_at": datetime.utcnow().isoformat()
        }
    ]

    if request.format == "csv":
        return export_companies_csv(companies)
    elif request.format == "excel":
        return export_companies_excel(companies)
    else:
        raise HTTPException(status_code=400, detail="Invalid format. Use 'csv' or 'excel'")


@router.post("/personnel")
async def export_personnel(
    request: ExportRequest,
    user: User = Depends(get_current_user_from_header)
):
    """
    Export personnel data to CSV or Excel.
    Protected route - requires authentication.
    """
    ***REMOVED*** Mock personnel data - in production, get from database
    personnel = [
        {
            "personnel_id": "pers-001",
            "name": "John Doe",
            "title": "Purchasing Manager",
            "company": "Grand Hotel Istanbul",
            "email": "john.doe@grandhotel.com",
            "phone": "+90 212 123 4567",
            "linkedin": "linkedin.com/in/johndoe",
            "department": "Purchasing",
            "source": "LinkedIn",
            "created_at": datetime.utcnow().isoformat()
        },
        {
            "personnel_id": "pers-002",
            "name": "Jane Smith",
            "title": "R&D Director",
            "company": "Bakery Co.",
            "email": "jane.smith@bakeryco.com",
            "phone": "+90 216 234 5678",
            "linkedin": "linkedin.com/in/janesmith",
            "department": "R&D",
            "source": "LinkedIn",
            "created_at": datetime.utcnow().isoformat()
        }
    ]

    if request.format == "csv":
        return export_personnel_csv(personnel)
    elif request.format == "excel":
        return export_personnel_excel(personnel)
    else:
        raise HTTPException(status_code=400, detail="Invalid format. Use 'csv' or 'excel'")


@router.post("/combined")
async def export_combined(
    request: ExportRequest,
    user: User = Depends(get_current_user_from_header)
):
    """
    Export both company and personnel data.
    Protected route - requires authentication.
    """
    ***REMOVED*** This would combine both datasets
    ***REMOVED*** For now, return companies as combined export
    companies = [
        {
            "company_id": "comp-001",
            "name": "Grand Hotel Istanbul",
            "address": "Taksim, Istanbul",
            "website": "www.grandhotel.com",
            "phone": "+90 212 123 4567",
            "email": "info@grandhotel.com",
            "source": "Official Source",
            "status": "Verified",
            "created_at": datetime.utcnow().isoformat()
        }
    ]

    personnel = [
        {
            "personnel_id": "pers-001",
            "name": "John Doe",
            "title": "Purchasing Manager",
            "company": "Grand Hotel Istanbul",
            "email": "john.doe@grandhotel.com",
            "phone": "+90 212 123 4567",
            "linkedin": "linkedin.com/in/johndoe",
            "department": "Purchasing",
            "source": "LinkedIn",
            "created_at": datetime.utcnow().isoformat()
        }
    ]

    if request.format == "csv":
        return export_combined_csv(companies, personnel)
    elif request.format == "excel":
        return export_combined_excel(companies, personnel)
    else:
        raise HTTPException(status_code=400, detail="Invalid format. Use 'csv' or 'excel'")


def export_companies_csv(companies: List[Dict]) -> StreamingResponse:
    """Export companies to CSV"""
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=[
        "company_id", "name", "address", "website", "phone", "email", 
        "source", "status", "created_at"
    ])
    writer.writeheader()
    writer.writerows(companies)
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=companies_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"}
    )


def export_personnel_csv(personnel: List[Dict]) -> StreamingResponse:
    """Export personnel to CSV"""
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=[
        "personnel_id", "name", "title", "company", "email", "phone",
        "linkedin", "department", "source", "created_at"
    ])
    writer.writeheader()
    writer.writerows(personnel)
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=personnel_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"}
    )


def export_combined_csv(companies: List[Dict], personnel: List[Dict]) -> StreamingResponse:
    """Export combined data to CSV"""
    output = io.StringIO()
    
    ***REMOVED*** Write companies
    output.write("=== COMPANIES ===\n")
    writer = csv.DictWriter(output, fieldnames=[
        "company_id", "name", "address", "website", "phone", "email",
        "source", "status", "created_at"
    ])
    writer.writeheader()
    writer.writerows(companies)
    
    ***REMOVED*** Write personnel
    output.write("\n=== PERSONNEL ===\n")
    writer = csv.DictWriter(output, fieldnames=[
        "personnel_id", "name", "title", "company", "email", "phone",
        "linkedin", "department", "source", "created_at"
    ])
    writer.writeheader()
    writer.writerows(personnel)
    
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename=combined_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"}
    )


def export_companies_excel(companies: List[Dict]) -> StreamingResponse:
    """Export companies to Excel"""
    try:
        import openpyxl
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Companies"
        
        ***REMOVED*** Headers
        headers = ["Company ID", "Name", "Address", "Website", "Phone", "Email", "Source", "Status", "Created At"]
        ws.append(headers)
        
        ***REMOVED*** Data
        for company in companies:
            ws.append([
                company.get("company_id", ""),
                company.get("name", ""),
                company.get("address", ""),
                company.get("website", ""),
                company.get("phone", ""),
                company.get("email", ""),
                company.get("source", ""),
                company.get("status", ""),
                company.get("created_at", "")
            ])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=companies_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"}
        )
    except ImportError:
        ***REMOVED*** Fallback to CSV if openpyxl not available
        return export_companies_csv(companies)


def export_personnel_excel(personnel: List[Dict]) -> StreamingResponse:
    """Export personnel to Excel"""
    try:
        import openpyxl
        from openpyxl import Workbook
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Personnel"
        
        ***REMOVED*** Headers
        headers = ["Personnel ID", "Name", "Title", "Company", "Email", "Phone", "LinkedIn", "Department", "Source", "Created At"]
        ws.append(headers)
        
        ***REMOVED*** Data
        for person in personnel:
            ws.append([
                person.get("personnel_id", ""),
                person.get("name", ""),
                person.get("title", ""),
                person.get("company", ""),
                person.get("email", ""),
                person.get("phone", ""),
                person.get("linkedin", ""),
                person.get("department", ""),
                person.get("source", ""),
                person.get("created_at", "")
            ])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=personnel_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"}
        )
    except ImportError:
        ***REMOVED*** Fallback to CSV if openpyxl not available
        return export_personnel_csv(personnel)


def export_combined_excel(companies: List[Dict], personnel: List[Dict]) -> StreamingResponse:
    """Export combined data to Excel"""
    try:
        import openpyxl
        from openpyxl import Workbook
        
        wb = Workbook()
        
        ***REMOVED*** Companies sheet
        ws_companies = wb.active
        ws_companies.title = "Companies"
        headers = ["Company ID", "Name", "Address", "Website", "Phone", "Email", "Source", "Status", "Created At"]
        ws_companies.append(headers)
        for company in companies:
            ws_companies.append([
                company.get("company_id", ""),
                company.get("name", ""),
                company.get("address", ""),
                company.get("website", ""),
                company.get("phone", ""),
                company.get("email", ""),
                company.get("source", ""),
                company.get("status", ""),
                company.get("created_at", "")
            ])
        
        ***REMOVED*** Personnel sheet
        ws_personnel = wb.create_sheet("Personnel")
        headers = ["Personnel ID", "Name", "Title", "Company", "Email", "Phone", "LinkedIn", "Department", "Source", "Created At"]
        ws_personnel.append(headers)
        for person in personnel:
            ws_personnel.append([
                person.get("personnel_id", ""),
                person.get("name", ""),
                person.get("title", ""),
                person.get("company", ""),
                person.get("email", ""),
                person.get("phone", ""),
                person.get("linkedin", ""),
                person.get("department", ""),
                person.get("source", ""),
                person.get("created_at", "")
            ])
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=combined_export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"}
        )
    except ImportError:
        ***REMOVED*** Fallback to CSV if openpyxl not available
        return export_combined_csv(companies, personnel)


@router.get("/health")
def health():
    return {
        "status": "ok",
        "domain": "export",
        "formats": ["csv", "excel"]
    }

